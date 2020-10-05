from matplotlib import rc
import openpyxl as xl
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FormatStrFormatter


def CatmullRomSpline(P0, P1, P2, P3, nPoints=100):
    """
    P0, P1, P2, and P3 should be (x,y) point pairs that define the
    Catmull-Rom spline.
    nPoints is the number of points to include in this curve segment.
    """
    # Convert the points to numpy so that we can do array multiplication
    P0, P1, P2, P3 = map(np.array, [P0, P1, P2, P3])

    # Calculate t0 to t4
    alpha = 0.5

    def tj(ti, Pi, Pj):
        xi, yi = Pi
        xj, yj = Pj
        return (((xj - xi)**2 + (yj - yi)**2)**0.5)**alpha + ti

    t0 = 0
    t1 = tj(t0, P0, P1)
    t2 = tj(t1, P1, P2)
    t3 = tj(t2, P2, P3)

    # Only calculate points between P1 and P2
    t = np.linspace(t1, t2, nPoints)

    # Reshape so that we can multiply by the points P0 to P3
    # and get a point for each value of t.
    t = t.reshape(len(t), 1)

    A1 = (t1 - t) / (t1 - t0) * P0 + (t - t0) / (t1 - t0) * P1
    A2 = (t2 - t) / (t2 - t1) * P1 + (t - t1) / (t2 - t1) * P2
    A3 = (t3 - t) / (t3 - t2) * P2 + (t - t2) / (t3 - t2) * P3

    B1 = (t2 - t) / (t2 - t0) * A1 + (t - t0) / (t2 - t0) * A2
    B2 = (t3 - t) / (t3 - t1) * A2 + (t - t1) / (t3 - t1) * A3

    C = (t2 - t) / (t2 - t1) * B1 + (t - t1) / (t2 - t1) * B2
    return C


def CatmullRomChain(P):
    """
    Calculate Catmull Rom for a chain of points and return the combined curve.
    """
    sz = len(P)

    # The curve C will contain an array of (x,y) points.
    C = []
    for i in range(sz - 3):
        c = CatmullRomSpline(P[i], P[i + 1], P[i + 2], P[i + 3])
        C.extend(c)

    return C


plt.close('all')

x_axes = {'labelpad': 12, 'fontsize': 24}
y_axes = {'labelpad': 24, 'fontsize': 24}

title_format = {'pad': 35, 'fontsize': 40}

rc('font', **{'family': 'serif', 'serif': ['Arial']})

dispColor = 'tab:blue'

x_step = 5
y_step = 5

# define axes limits
# x_lim = [
#     (18, 38),
#     (22, 42),
#     (20, 40),
#     (12, 32),
#     (18, 38),
#     (16, 36),
#     (24, 44),
#     (14, 34),
#     (14, 34),
#     (18, 38)
# ]
x_lim = [(20, 40), (25, 45), (20, 40), (10, 30), (20, 40), (15, 35), (25, 45),
         (15, 35), (15, 35), (20, 40)]
y_lim = [(0, 2), (0, 2), (0, 2), (0, 2)]
sec_y_lim = [(0.7, 0.8), (1.95, 2.15), (2.6, 3), (-1.2, -0.8)]
plane_labels = [
    r'$(101)$',
    r'$(11\bar1)$',
    r'$(\bar111)$',
    r'$(2\bar10)$',
    r'$(210)$',
    r'$(0\bar12)$',
    r'$(012)$',
    r'$(1\bar21)$',
    r'$(\bar230)$',
    r'$(0\bar32)$',
]

site_labels = ['Cu', 'Bi', 'W', 'O']
stp = [1, 2, 3, 4]
# stp = [3, 4]

# ptp = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # planes to plot
# ptp = [6, 1, 3, 8, 4]  # planes to plot
ptp = [4]  # planes to plot

# main figure/subplots
fig, axes = plt.subplots(figsize=(3.6, 8.2),
                         ncols=len(ptp),
                         nrows=len(stp),
                         sharex='col',
                         sharey='row')

if type(axes) is not np.ndarray:
    # single pane
    axes = [[axes]]
else:
    axes = axes.tolist()
    if type(axes[0]) != list:
        # single column
        axes = [[ax] for ax in axes]

# wrapper axes object to center labels (including secondary y-axis)
ax_big = fig.add_subplot(111, frameon=False)
ax_big.set_xlabel("z (Å)", **x_axes)
ax_big.set_ylabel("Displacement (Å)", **y_axes, c=dispColor)
ax_big.tick_params(labelcolor='none',
                   top=False,
                   bottom=False,
                   left=False,
                   right=False)
ax_big_sec = ax_big.secondary_yaxis('right')
ax_big_sec.set_frame_on(False)
ax_big_sec.set_ylabel("Net Charge (e)", **y_axes, c='k')
ax_big_sec.tick_params(labelcolor='none',
                       top=False,
                       bottom=False,
                       left=False,
                       right=False)

# secondary axes array
sec_axes = [[[] for j in range(len(ptp))] for i in range(len(stp))]

# set surface labels
for i in range(len(ptp)):
    axes[0][i].set_title(plane_labels[ptp[i] - 1], **title_format)

# set x-axis
l = len(stp) - 1
for i in range(len(ptp)):
    axes[l][i].set_xlim(x_lim[ptp[i] - 1])
    axes[l][i].set_xticks(np.linspace(*(x_lim[ptp[i] - 1] + (x_step, ))))
    axes[l][i].tick_params(axis='x', labelsize=12)

# set ticks to all axes
for i in range(len(stp)):

    # set primary y-axis
    axes[i][0].set_ylim(y_lim[stp[i] - 1])
    axes[i][0].set_yticks(np.linspace(*(y_lim[stp[i] - 1] + (y_step, ))))
    axes[i][0].yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
    x = axes[i][0].get_xlim()[0]
    y = axes[i][0].get_ylim()[1]
    axes[i][0].text(1.05 * x, .8 * y, site_labels[stp[i] - 1], fontsize=16)
    axes[i][0].tick_params(axis='y', labelsize=12)

    # set secondary y-axis
    for j in range(len(ptp)):
        sec_axes[i][j] = axes[i][j].twinx()
        sec_axes[i][j].set_ylim(sec_y_lim[stp[i] - 1])
        sec_axes[i][j].set_yticks(
            np.linspace(*(sec_y_lim[stp[i] - 1] + (y_step, ))))
        if j != len(ptp) - 1:
            sec_axes[i][j].tick_params(labelright=False)
        else:
            sec_axes[i][j].tick_params(axis='y', labelsize=12)
            sec_axes[i][j].yaxis.set_major_formatter(
                FormatStrFormatter('%.2f'))

# set data
wb = xl.load_workbook(r'C:\Users\edanb\Desktop\data.xlsx', data_only=True)

c = [1, 4, 5]
for j in range(len(ptp)):

    # get sheet
    sheet = wb[str(ptp[j])]

    for i in range(len(stp)):

        # get column letters
        xc = get_column_letter(c[0] + 6 * (stp[i] - 1))
        yc1 = get_column_letter(c[1] + 6 * (stp[i] - 1))
        yc2 = get_column_letter(c[2] + 6 * (stp[i] - 1))

        x = [cell.value for cell in sheet[xc:xc] if cell.value is not None]
        yd = [cell.value for cell in sheet[yc1:yc1] if cell.value is not None]
        yb = [cell.value for cell in sheet[yc2:yc2] if cell.value is not None]

        C1 = CatmullRomChain([[i, j] for i, j in zip(x, yd)])
        C2 = CatmullRomChain([[i, j] for i, j in zip(x, yb)])

        x1, y1 = zip(*C1)
        # x1 = (x[0],) + x1 + (x[-1],)
        # y1 = (yd[0],) + y1 + (yd[-1],)

        x2, y2 = zip(*C2)
        # x2 = (x[0],) + x2 + (x[-1],)
        # y2 = (yb[0],) + y2 + (yb[-1],)

        axes[i][j].plot(x1, y1, dispColor, lw=2.0)
        sec_axes[i][j].plot(x2, y2, 'k', ls=':', lw=0.75)

# plt.tight_layout()
plt.subplots_adjust(top=0.87,
                    bottom=0.1,
                    left=0.281,
                    right=0.732,
                    hspace=0.32,
                    wspace=0.2)

plt.savefig(r'C:\Users\edanb\Desktop\dispFig.jpeg', dpi=600, format='jpeg')

plt.show()
